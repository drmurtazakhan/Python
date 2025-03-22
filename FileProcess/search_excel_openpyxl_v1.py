import os
import openpyxl

# Clearing the Screen
os.system('cls')

# Provide the path to your Excel file
fileName = 'C:/0MAK/RW/PythonKhan/DataFiles/CSV/ID_Marks2.xlsx'

# Read the Excel file
theFile =  openpyxl.load_workbook(fileName)

allSheetNames = theFile.sheetnames

print("All sheet names {} " .format(theFile.sheetnames))

    
def find_specific_cell():
    for row in range(1, currentSheet.max_row + 1):
        for column in "A":  # Here you can add or reduce the columns
            cell_name = "{}{}".format(column, row)
            if currentSheet[cell_name].value == 201631:
                #print("{1} cell is located on {0}" .format(cell_name, currentSheet[cell_name].value))
                print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))
                return cell_name


for sheet in allSheetNames:
    print("Current sheet name is {}" .format(sheet))
    currentSheet = theFile[sheet]
    specificCellLetter = (find_specific_cell())
    



   

    